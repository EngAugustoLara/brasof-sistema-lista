
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema Lista BraSoft – Preenchimento Diário (v1.5.2)
"""
import json, os, re, sys, unicodedata
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

APP_TITLE = "Sistema Lista BraSoft – Preenchimento Diário"
MAP_FILENAME = "sku_map.json"

def app_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent

def strip_accents(s: str) -> str:
    try:
        return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    except Exception:
        return s

def normalize_token(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = strip_accents(s).lower()
    s = re.sub(r'\s+', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def default_map_dict() -> dict:
    return {
        "priorities": ["PR","R70","ENC","ENC_CAPA","XUXAO","BB","RAMPA"],
        "categories": {},
        "special_rules": {"enc_capa_category": "ENC_CAPA", "enc_base_category": "ENC"}
    }

def load_map_file(path: Path) -> dict:
    if not path.exists():
        return default_map_dict()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        data.setdefault("priorities", default_map_dict()["priorities"])
        data.setdefault("categories", {})
        data.setdefault("special_rules", default_map_dict()["special_rules"])
        for cat, meta in data["categories"].items():
            meta["aliases"] = sorted(set([normalize_token(a) for a in meta.get("aliases", [])]))
            meta.setdefault("output_format", "numeric")
        return data
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao ler {path.name}:\n{e}")
        return default_map_dict()

def save_map_file(path: Path, data: dict):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        for cat, meta in data.get("categories", {}).items():
            meta["aliases"] = sorted(set([normalize_token(a) for a in meta.get("aliases", [])]))
            meta.setdefault("output_format", "numeric")
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar mapa em {path}:\n{e}")

# regex
QTD_PATTERNS = [
    re.compile(r'Quantity:\s*(\d+)', re.IGNORECASE),
    re.compile(r'"Quantity"\s*[:=]\s*"?(\d+)"?', re.IGNORECASE),
    re.compile(r'\bqty\s*[:=]\s*(\d+)', re.IGNORECASE),
    re.compile(r'\bquantity\s*[:=]\s*(\d+)', re.IGNORECASE),
    re.compile(r'^\s*\[(\d+)\]\s', re.IGNORECASE),
]
SKU_IN_PRODUCT_INFO = re.compile(r'SKU Reference No\.\s*:\s*([A-Za-z0-9_\-\. ]+)', re.IGNORECASE)
VAR_IN_PRODUCT_INFO = re.compile(r'Variation Name\s*:\s*([^;\\n]+)', re.IGNORECASE)
BLOCK_RE = re.compile(r'\[\d+\][^\[]+', re.IGNORECASE | re.DOTALL)

def extract_quantity_from_text(text: str, fallback=None):
    if isinstance(text, str):
        for pat in QTD_PATTERNS:
            m = pat.search(text)
            if m:
                try:
                    return int(m.group(1))
                except Exception:
                    continue
    return 1 if fallback is None else fallback

def extract_sku_and_var_from_text(text: str):
    sku_raw = ""
    variation = ""
    if not isinstance(text, str) or text.strip() == "":
        return sku_raw, variation
    m = SKU_IN_PRODUCT_INFO.search(text)
    if m:
        sku_raw = m.group(1).strip()
    m2 = VAR_IN_PRODUCT_INFO.search(text)
    if m2:
        variation = m2.group(1).strip()
    return sku_raw, variation

def quant_kit_from_variation(variation: str) -> int:
    if not isinstance(variation, str) or not variation.strip():
        return 1
    m = re.search(r'\b(\d+)\b', variation)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return 1
    return 1

def compute_category_from_variation(initial_category, variation, enc_capa_cat, enc_base_cat):
    v = variation or ""
    v_norm = normalize_token(v)
    has_encosto = "encosto" in v_norm
    has_capa_extra = "capaextra" in v_norm
    has_com_capa = "comcapa" in v_norm or "+capa" in v_norm
    if has_encosto and has_capa_extra:
        return enc_capa_cat
    if has_encosto and has_com_capa and initial_category != enc_capa_cat:
        return enc_base_cat
    return initial_category

def compute_outputs_single_item(sku_raw, variation, qty_from_block, cfg: dict):
    kit_qty = quant_kit_from_variation(variation)
    kits_purchased = qty_from_block if qty_from_block else 1
    unidades = kit_qty * kits_purchased
    sku_norm = normalize_token(sku_raw)
    category = ""
    for cat, meta in cfg["categories"].items():
        if sku_norm in meta.get("aliases", []):
            category = cat
            break
    enc_capa_cat = cfg.get("special_rules", {}).get("enc_capa_category", "ENC_CAPA")
    enc_base_cat = cfg.get("special_rules", {}).get("enc_base_category", "ENC")
    category = compute_category_from_variation(category, variation, enc_capa_cat, enc_base_cat)
    outputs = {}
    if category:
        if category == enc_capa_cat and "capaextra" in normalize_token(variation or ""):
            outputs[category] = {"type": "enc_capa", "value": kits_purchased}
        else:
            outputs[category] = {"type": "numeric", "value": unidades}
    diag = {
        "sku_raw": sku_raw or "", "sku_norm": sku_norm, "category": category,
        "kit_qty": kit_qty, "kits_purchased": kits_purchased, "unidades": unidades,
        "variation_seen": variation or ""
    }
    return outputs, diag

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x660")
        self.resizable(True, True)

        self.map_path = app_dir() / MAP_FILENAME
        if not self.map_path.exists():
            save_map_file(self.map_path, default_map_dict())
        self.cfg = load_map_file(self.map_path)

        self.orders_path = None
        self._build_ui()
        self._refresh_list()

    def _build_ui(self):
        pad = 8
        menubar = tk.Menu(self)
        m_map = tk.Menu(menubar, tearoff=0)
        m_map.add_command(label="Abrir pasta do mapa", command=self.open_map_folder)
        m_map.add_command(label="Carregar mapa...", command=self.load_map_from_file)
        m_map.add_command(label="Salvar mapa...", command=self.save_map_to_file)
        menubar.add_cascade(label="Mapa", menu=m_map)
        self.config(menu=menubar)

        top = ttk.LabelFrame(self, text=f"1) Selecione a planilha da Shopee — Mapa: {self.map_path}")
        top.pack(fill="x", padx=pad, pady=(pad, 0))
        self.lbl_file = ttk.Label(top, text="Nenhum arquivo selecionado")
        self.lbl_file.pack(side="left", padx=pad, pady=pad)
        ttk.Button(top, text="Selecionar planilha...", command=self.pick_file).pack(side="right", padx=pad, pady=pad)

        mid = ttk.LabelFrame(self, text="2) Colunas/Produtos e seus SKUs (aliases)")
        mid.pack(fill="both", expand=True, padx=pad, pady=pad)

        left = ttk.Frame(mid); left.pack(side="left", fill="both", expand=True, padx=pad, pady=pad)
        ttk.Label(left, text="Colunas/Produtos").pack(anchor="w")
        self.list_categories = tk.Listbox(left, height=15); self.list_categories.pack(fill="both", expand=True)
        self.list_categories.bind("<<ListboxSelect>>", self.on_select_category)

        btns = ttk.Frame(left); btns.pack(fill="x", pady=(pad, 0))
        ttk.Button(btns, text="+ Adicionar novo SKU/Coluna", command=self.add_category_dialog).pack(side="left", padx=2)
        ttk.Button(btns, text="✎ Renomear", command=self.edit_category_dialog).pack(side="left", padx=2)
        ttk.Button(btns, text="🗑️ Remover", command=self.remove_category).pack(side="left", padx=2)

        right = ttk.Frame(mid); right.pack(side="right", fill="both", expand=True, padx=pad, pady=pad)
        ttk.Label(right, text="Aliases (formas como o SKU aparece em 'SKU Reference No.')").pack(anchor="w")
        self.txt_aliases = tk.Text(right, height=8); self.txt_aliases.pack(fill="both", expand=True)

        small = ttk.Frame(right); small.pack(fill="x", pady=(pad, 0))
        ttk.Button(small, text="Salvar alterações", command=self.save_aliases).pack(side="left")
        self.info_hint = ttk.Label(small, text="Dica: separe por vírgulas. Ex.: '2Promo, 4Promo, 2PromoNovo'")
        self.info_hint.pack(side="left", padx=10)

        bottom = ttk.LabelFrame(self, text="3) Gerar planilha"); bottom.pack(fill="x", padx=pad, pady=pad)
        self.chk_diag = tk.BooleanVar(value=True)
        ttk.Checkbutton(bottom, text="Salvar aba 'Diagnostico'", variable=self.chk_diag).pack(side="left", padx=pad)
        ttk.Button(bottom, text="Gerar Planilha", command=self.run_process).pack(side="right", padx=pad, pady=pad)

        self.status = ttk.Label(self, text=f"Pronto. (Mapa: {self.map_path})", anchor="w")
        self.status.pack(fill="x", padx=pad, pady=(0, pad))

    # menu
    def open_map_folder(self):
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(self.map_path.parent))
            elif sys.platform == "darwin":
                os.system(f'open "{self.map_path.parent}"')
            else:
                os.system(f'xdg-open "{self.map_path.parent}"')
        except Exception as e:
            messagebox.showerror("Erro", f"Não consegui abrir a pasta:\n{e}")

    def load_map_from_file(self):
        f = filedialog.askopenfilename(title="Selecione um sku_map.json", filetypes=[("JSON", "*.json")])
        if not f: return
        p = Path(f)
        self.cfg = load_map_file(p)
        self.map_path = p
        self._refresh_list()
        self.status.config(text=f"Mapa carregado de: {p}")

    def save_map_to_file(self):
        f = filedialog.asksaveasfilename(title="Salvar sku_map.json", defaultextension=".json", initialfile="sku_map.json", filetypes=[("JSON", "*.json")])
        if not f: return
        p = Path(f)
        save_map_file(p, self.cfg)
        messagebox.showinfo("OK", f"Mapa salvo em:\n{p}")

    # ui
    def pick_file(self):
        f = filedialog.askopenfilename(title="Selecione a planilha da Shopee", filetypes=[("Excel", "*.xlsx")])
        if f:
            self.orders_path = Path(f)
            self.lbl_file.config(text=str(self.orders_path))

    def _refresh_list(self):
        self.list_categories.delete(0, tk.END)
        cats = sorted(self.cfg["categories"].keys(), key=str.lower)
        prios = [c for c in self.cfg.get("priorities", []) if c in cats]
        others = [c for c in cats if c not in prios]
        cats = prios + others
        for c in cats:
            self.list_categories.insert(tk.END, c)
        self.txt_aliases.delete("1.0", tk.END)

    def on_select_category(self, event=None):
        idxs = self.list_categories.curselection()
        if not idxs:
            self.txt_aliases.delete("1.0", tk.END); return
        cat = self.list_categories.get(idxs[0])
        aliases = self.cfg["categories"].get(cat, {}).get("aliases", [])
        self.txt_aliases.delete("1.0", tk.END)
        self.txt_aliases.insert("1.0", ", ".join(aliases))

    def add_category_dialog(self):
        name = simpledialog.askstring("Nova coluna/produto", "Nome da coluna/produto (ex.: PR, R70, ENC, PILLOWTOP):", parent=self)
        if not name: return
        name = name.strip()
        if not name: return
        if name in self.cfg["categories"]:
            messagebox.showinfo("Aviso", f"A coluna '{name}' já existe."); return
        aliases_str = simpledialog.askstring("Aliases", "Digite as variações do SKU (como em 'SKU Reference No.')\nSepare por vírgulas.", parent=self)
        aliases = []
        if aliases_str:
            aliases = [normalize_token(x) for x in aliases_str.split(",") if x.strip()]
        self.cfg["categories"][name] = {"aliases": sorted(set(aliases)), "output_format": "numeric"}
        save_map_file(self.map_path, self.cfg)
        self._refresh_list()
        messagebox.showinfo("OK", f"Coluna '{name}' criada.")

    def edit_category_dialog(self):
        idxs = self.list_categories.curselection()
        if not idxs:
            messagebox.showinfo("Info", "Selecione uma coluna para renomear."); return
        old = self.list_categories.get(idxs[0])
        new = simpledialog.askstring("Renomear coluna", f"Novo nome para '{old}':", initialvalue=old, parent=self)
        if not new: return
        new = new.strip()
        if new == "" or new == old: return
        if new in self.cfg["categories"]:
            messagebox.showerror("Erro", f"Já existe a coluna '{new}'."); return
        self.cfg["categories"][new] = self.cfg["categories"].pop(old)
        prios = self.cfg.get("priorities", [])
        self.cfg["priorities"] = [new if p == old else p for p in prios]
        save_map_file(self.map_path, self.cfg)
        self._refresh_list()
        messagebox.showinfo("OK", f"Renomeada para '{new}'.")

    def remove_category(self):
        idxs = self.list_categories.curselection()
        if not idxs:
            messagebox.showinfo("Info", "Selecione uma coluna para remover."); return
        cat = self.list_categories.get(idxs[0])
        if messagebox.askyesno("Confirmar", f"Remover a coluna '{cat}'?"):
            self.cfg["categories"].pop(cat, None)
            self.cfg["priorities"] = [p for p in self.cfg.get("priorities", []) if p != cat]
            save_map_file(self.map_path, self.cfg)
            self._refresh_list()

    def save_aliases(self):
        idxs = self.list_categories.curselection()
        if not idxs:
            messagebox.showinfo("Info", "Selecione uma coluna para salvar os aliases."); return
        cat = self.list_categories.get(idxs[0])
        aliases_str = self.txt_aliases.get("1.0", tk.END).strip()
        aliases = [normalize_token(x) for x in aliases_str.split(",") if x.strip()]
        if cat not in self.cfg["categories"]:
            self.cfg["categories"][cat] = {"aliases": [], "output_format": "numeric"}
        self.cfg["categories"][cat]["aliases"] = sorted(set(aliases))
        save_map_file(self.map_path, self.cfg)
        messagebox.showinfo("OK", "Aliases salvos!")

    def run_process(self):
        if not self.orders_path or not self.orders_path.exists():
            messagebox.showerror("Erro", "Selecione a planilha da Shopee (orders.xlsx)."); return
        try:
            df = pd.read_excel(self.orders_path)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao abrir Excel:\n{e}"); return

        product_info_col = "product_info" if "product_info" in df.columns else None
        order_col = "order_sn" if "order_sn" in df.columns else None
        if product_info_col is None:
            messagebox.showerror("Erro", "Não encontrei a coluna 'product_info'."); return

        categories = list(self.cfg["categories"].keys())
        prios = [c for c in self.cfg.get("priorities", []) if c in categories]
        remaining = [c for c in sorted(categories, key=str.lower) if c not in prios]
        out_cols = prios + remaining

        detail_rows, diag_rows = [], []

        for _, row in df.iterrows():
            order_sn = row.get(order_col, '')
            product_info = row.get(product_info_col, '')

            blocks = BLOCK_RE.findall(product_info) if isinstance(product_info, str) else []
            if not blocks:
                blocks = [product_info] if isinstance(product_info, str) else []

            accum_numeric = {}
            accum_enc_capa = 0
            per_block_variations, per_block_sku = [], []

            for blk in blocks:
                sku_raw, variation = extract_sku_and_var_from_text(blk)
                qty_blk = extract_quantity_from_text(blk, fallback=1)
                outputs, diag = compute_outputs_single_item(sku_raw, variation, qty_blk, self.cfg)
                per_block_variations.append(diag["variation_seen"])
                per_block_sku.append(diag["sku_raw"])
                diag_rows.append({"order_sn": order_sn, **diag})
                for col, meta in outputs.items():
                    if meta["type"] == "numeric":
                        accum_numeric[col] = accum_numeric.get(col, 0) + meta["value"]
                    elif meta["type"] == "enc_capa":
                        accum_enc_capa += meta["value"]

            base = {'order_sn': order_sn,
                    'SKU Reference No.': "; ".join([s for s in per_block_sku if s]),
                    'Variation Name': "; ".join([v for v in per_block_variations if v]),
                    'product_info': product_info if isinstance(product_info, str) else str(product_info)}
            for c in out_cols: base[c] = ""
            for c, val in accum_numeric.items(): base[c] = val
            if accum_enc_capa > 0: base["ENC_CAPA"] = f"{accum_enc_capa} + C"
            detail_rows.append(base)

        df_det = pd.DataFrame(detail_rows, columns=['order_sn', 'SKU Reference No.', 'Variation Name', 'product_info'] + out_cols)

        def sum_or_blank(series):
            try:
                numeric = pd.to_numeric(series, errors='coerce')
                s = numeric.sum(min_count=1)
                return "" if pd.isna(s) else int(s)
            except Exception:
                return ""

        agg_map = {}
        for c in out_cols:
            if c == "ENC_CAPA":
                agg_map[c] = lambda s: sum([int(str(x).split()[0]) for x in s if str(x).strip()])
            else:
                agg_map[c] = sum_or_blank
        resumo = df_det.groupby('order_sn', as_index=False).agg(agg_map)

        total_row = {'order_sn': 'TOTAL'}
        for c in out_cols:
            if c == "ENC_CAPA":
                nums = []
                for x in df_det[c].tolist():
                    try:
                        n = int(str(x).split()[0]); nums.append(n)
                    except Exception: pass
                total_row[c] = (f"{sum(nums)} + C") if nums else ""
            else:
                total_row[c] = sum_or_blank(df_det[c])
        resumo = pd.concat([resumo, pd.DataFrame([total_row])], ignore_index=True)

        out_path = self.orders_path.with_name("preenchido.xlsx")
        try:
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                df_det.to_excel(writer, sheet_name='ItensDetalhados', index=False)
                resumo.to_excel(writer, sheet_name='Resumo', index=False)
                pd.DataFrame(diag_rows).to_excel(writer, sheet_name='Diagnostico', index=False)

                wb = writer.book; ws_det = wb['ItensDetalhados']
                yellow = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
                header_row = 1; data_start = header_row + 1
                for r in range(data_start, data_start + len(df_det)):
                    start_idx = 5
                    if all((ws_det.cell(row=r, column=start_idx + i).value in (None, "", 0)) for i in range(len(out_cols))):
                        for c in range(1, ws_det.max_column + 1):
                            ws_det.cell(row=r, column=c).fill = yellow

                def autosize(ws):
                    widths = {}
                    for row in ws.iter_rows():
                        for cell in row:
                            val = cell.value
                            length = len(str(val)) if val is not None else 0
                            col = cell.column
                            widths[col] = max(widths.get(col, 0), length)
                    for i, w in widths.items():
                        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)

                autosize(ws_det); autosize(wb['Resumo']); autosize(wb['Diagnostico'])
                writer._save()

            self.status.config(text=f"Gerado: {out_path.name}")
            messagebox.showinfo("Concluído", f"Arquivo gerado:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao salvar Excel:\n{e}")

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
